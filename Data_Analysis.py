#!/usr/bin/env python
# coding: utf-8

# In[1]:


import plotly as py
import numpy as np
import math as m
from scipy.odr import *
py.offline.init_notebook_mode(connected=True)
import plotly.graph_objs as go
import sympy as sp


# # Media

# In[2]:


def mean(x:list):
    return sum(x)/len(x)


# # Deviazione standard $\sigma$

# In[3]:


def std(x:list):
    media = mean(x)
    return m.sqrt((sum([(i-media)**2 for i in x]))/(len(x)-1))


# # Deviazione standard della media $\sigma_{\bar{x}}$

# In[4]:


def std_mean(x:list):
    media = mean(x)
    return m.sqrt((sum([(i-media)**2 for i in x]))/(len(x)-1))/m.sqrt(len(x))


# In[ ]:





# # Compatibilità

# In[5]:


def compatibilità(x, errore_x, y, errore_y, docs:bool = False):
    """Ritorna il valore compatibilità o una stringa a seconda che docs sia False o True; Compatibilità di default
    """
    num = abs(x-y)
    den = m.sqrt(errore_x**2 + errore_y**2)
    r = num/den
    output = ""
    if r >= 0 and r < 1:
        output += f"Buona, r = {round(r,3)}\n"
    elif r > 1 and r < 2:
        output += f"Sufficiente r = {round(r,3)}\n"
    elif r > 2 and r < 3:
        output += f"Scarsa r = {round(r,3)}\n"
    else:
        output += f"Incompatibilità r = {round(r,3)}\n"
    if docs:
        return output
    return r


# In[6]:


def compatibilità_list(x:list, errore_x:list, docs:bool = False):
    """Ritorna una stringa
    """
    output = ""
    for i, val1 in enumerate(x):
        for j, val2 in enumerate(x):
            if i == j or i-j > 0:
                pass
            else:
                output += f"La compatibilità tra l'elemento {i+1} e {j+1} è:"
                comp = compatibilità(val1,errore_x[i], val2,errore_x[j], True)
                output += comp
                output += " "
                print("La compatibilità tra l'elemento", {i+1}, "e ", {j+1}, " è:", comp)
    # while i < len(x)-1:
    #     output += f"La compatibilità tra l'elemento {i+1} e {i+2} è: \n"
    #     output += compatibilità(x[i],errore_x[i], x[i+1],errore_x[i+1])
    #     i += 1
    # while j < len(x)-2:
    #     output += f"La compatibilità tra l'elemento {j+1} e {j+2} è: \n"
    #     output += compatibilità(x[j], errore_x[j], x[j+2], errore_x[j+2])
    #     j += 1
    if docs: 
        return output


# 
# # Media Ponderata

# In[7]:


def media_ponderata(x:list, errore_x:list, docs:bool = False):
    """Ritorna un tuple: media, errore_media, stringa per il documento
    """
    num = 0
    den = 0
    output = ""
    for i in range(0,len(x)):
        num += x[i]/pow(errore_x[i],2)
        den += 1/pow(errore_x[i],2)
    media = num/den
    errore_media = 1/m.sqrt(den)
    if len(x) == 2:
        r = compatibilità(x[0],errore_x[0], x[1], errore_x[1])
        if r <3 :
            output += f"Poiché c'è compatibilità ne facciamo la media ponderata:\n"
    output += f"La media ponderata vale: {media}\n"
    output += f"La sue incertezza è: {errore_media}\n"
    print("La media ponderata vale: ", media)
    print("La sua incertezza è: ", errore_media)
    if docs:
        return output
    return media, errore_media


# # Coefficiente di Pearson

# In[8]:


def coefficiente_Pearson(x:list, y:list, docs:bool = False):
    sum1, sum2, sum_quadro_1, sum_quadro_2 = 0, 0, 0, 0
    output = ""
    for i in range(len(x)):
        sum1 += x[i]
    media_x = sum1/len(x)
    for j in range(len(y)):
        sum2 += y[j]
    media_y = sum2/len(y)
    for k in range(len(x)):
        sum_quadro_1 += pow(x[k]-media_x,2) 
        sum_quadro_2 += pow(y[k]-media_y,2)
    numeratore = 0
    denominatore = m.sqrt(sum_quadro_1)*m.sqrt(sum_quadro_2)
    for l in range(len(x)):
        numeratore += (x[l]-media_x)*(y[l]-media_y)
    rho = numeratore/denominatore
    output += f"L'indice di correlazione di Pearson vale: $\rho = $ {rho}\n"
    print("L'indice di correlazione di Pearson vale: rho = ", rho)
    if docs:
        return output
    return rho


# # Covarianza

# In[9]:


def covarianza(x:list, y:list):
    N = len(x)
    N_covariance = 0
    x_mean = sum(x)/N
    y_mean = sum(y)/N
    for i in range(N):
        N_covariance += (x[i] - x_mean)*(y[i] - y_mean)
    return N_covariance/N


# # INTERPOLAZIONE LINEARE  $y = a + bx $
# 

# In[10]:


def interpolazione(y:list, errore_y:list, x:list, errore_x = None, docs:bool = False):
    uno_sigma_quadro1, x_su_sigma_quadro1, x_per_y_su_sigma_quadro1, y_su_sigma_quadro1, x_quadro_su_sigma_quadro1 = 0, 0, 0, 0, 0

    for i in range (0, len(x)):
        uno_sigma_quadro1 += 1/pow(errore_y[i],2)
        x_su_sigma_quadro1 += x[i]/pow(errore_y[i],2)    
        x_quadro_su_sigma_quadro1  += pow(x[i],2)/pow(errore_y[i],2) 
        x_per_y_su_sigma_quadro1 += x[i]*y[i]/pow(errore_y[i],2)
        y_su_sigma_quadro1 += y[i]/pow(errore_y[i],2)

    delta = uno_sigma_quadro1*x_quadro_su_sigma_quadro1-pow(x_su_sigma_quadro1,2)
    b = 1/delta * (uno_sigma_quadro1*x_per_y_su_sigma_quadro1 - x_su_sigma_quadro1*y_su_sigma_quadro1)
    a = 1/delta * (x_quadro_su_sigma_quadro1*y_su_sigma_quadro1 - x_su_sigma_quadro1*x_per_y_su_sigma_quadro1)
    output = ""
    print("Il coefficiente angolare 'b' vale: ", b)
    errore_b = m.sqrt(1/delta * uno_sigma_quadro1)
    print("L'incertezza sul coefficiente angolare vale: ", errore_b)
    errore_a = m.sqrt(1/delta * x_quadro_su_sigma_quadro1)
    print("L'intercetta 'a' della retta vale: ", a)
    print("L'incertezza sull'intercetta vale: ", errore_a)
    output += f"Il coefficiente angolare 'b' vale: {b}\n"
    output += f"L'incertezza sul coefficiente angolare vale:: {errore_b}\n"
    output += f"IL'intercetta 'a' della retta vale: {a}\n"
    output += f"L'incertezza sull'intercetta vale: {errore_a}\n"
    if docs:
        return output
    return b, errore_b, a, errore_a


# In[11]:


def interpolazione_no_errore(x:list, y:list, errore_y, docs:bool = False):
    N = len(y)
    x_quadri, x_singoli, y_singoli, x_per_y = 0, 0, 0, 0
    for i in range (0,len(y)):
        x_quadri += pow(x[i],2)
        x_singoli += x[i]
        y_singoli += y[i]
        x_per_y += y[i]*x[i]

    delta = N*x_quadri -pow(x_singoli,2)
    a = (x_quadri*y_singoli-x_singoli*x_per_y)/delta
    b = (N*x_per_y-x_singoli*y_singoli)/delta
    errore_a = errore_y*m.sqrt(x_quadri/delta)
    errore_b = errore_y*m.sqrt(N/delta)
    print("Questo è a, l'intercetta della retta: ", a,"\nQuesto il suo errore:", errore_a, "\nQuesto è b, il coefficiente angolare", b, "\nQuesto il suo errore:", errore_b)
    output = ""
    output += f"Questo è a, l'intercetta della retta: {a}\nQuesto il suo errore: {errore_a} \nQuesto è b, il coefficiente angolare {b} \nQuesto il suo errore: {errore_b}"
    if docs:
        return output
    return (b, errore_b, a, errore_a)


# 
# # Grafico

# In[12]:


def grafico(b, a, y, errore_y, x, errore_x = None):
    
    layout = go.Layout(
        title= "",
        yaxis=dict(
                title = "y"
        ),
        xaxis=dict(
            title= "x",
        )
    )
    traccia = go.Scatter(
        x = x,
        y= y,
        mode='markers',
        name= "Traccia",
        showlegend=True,
       # line = dict(
        #    shape='spline'
       # ),
        error_y=dict(
                type='data',
                array = errore_y,
        ),
        error_x=dict(
                type='data',
                array = errore_x,
        ),
    )
#     traccia2 = go.Scatter(
#         x = raggi_quadro2,
#         y= velocità2,
#         mode='markers',
#         name='$v_i, i \in [1,9]$',
#         showlegend=True,
#        # line = dict(
#         #    shape='spline'
#        # )

#             error_y=dict(
#                 type='data',
#                 array = errore_velocità2,
#             )
#     )
    t = np.linspace(0., max(x), 1000)
    lalla = a + b*t
    if a >= 0:
        retta = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = "Retta"
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    else:
        retta = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = "Retta"
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    fig = go.Figure(data=[traccia, retta], layout = layout)
    py.offline.iplot(fig)


# In[13]:


def grafico_relazione(b, a, y, errore_y, x, errore_x = None):
    
    layout = go.Layout(
        title= str(input("Inserisci il nome del grafico ") ),
        yaxis=dict(
                title = str("$"+str(input("Inserisci il titolo dell'asse y ") ) + " [" + str(input("Inserisci l'unità di misura ")) + "]" + "$")
        ),
        xaxis=dict(
            title= str("$" + str(input("Inserisci il titolo dell'asse x ") ) + " [" +str(input("Inserisci l'unità di misura ")) + "]" +"$"),
        )
    )
    i = int(input("Scegli di quanto vuoi arrotondare b: ") )
    j = int(input("Scegli di quanto vuoi arrotondare a: ") )
    traccia = go.Scatter(
        x = x,
        y= y,
        mode='markers',
        name= "$" + str(input("Inserisci la legenda ") ) +"$",
        showlegend=True,
       # line = dict(
        #    shape='spline'
       # ),
        error_y=dict(
                type='data',
                array = errore_y,
        ),
        error_x=dict(
                type='data',
                array = errore_x,
        ),
    )
#     traccia2 = go.Scatter(
#         x = raggi_quadro2,
#         y= velocità2,
#         mode='markers',
#         name='$v_i, i \in [1,9]$',
#         showlegend=True,
#        # line = dict(
#         #    shape='spline'
#        # )

#             error_y=dict(
#                 type='data',
#                 array = errore_velocità2,
#             )
#     )
    t = np.linspace(0., max(x), 1000)
    lalla = a + b*t
    if a >= 0:
        retta = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = str("$y = " + "(" + str(round(b,i))+ "x +" + str(round(a,j))+ ")" + str(input("Inserisci l'unità di misura ")) +"$" )
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    else:
        retta = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = str("$y = " + "("+str(round(b,i))+ "x" + str(round(a,j)) +")" + str(input("Inserisci l'unità di misura ")) +"$")
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    fig = go.Figure(data=[traccia, retta], layout = layout)
    py.offline.iplot(fig)


# In[ ]:





# In[14]:


#z = interpolazione(velocità2, errore_velocità2, raggi_quadro2, None)


# In[15]:


#grafico(z[0], z[2], velocità2, errore_velocità2, raggi_quadro2)


# In[16]:


def chi_square(b, a, y:list, errore_y:list, x:list, docs:bool = False):
    #vincoli = N-2 -- caso interpolazione
    chi_quadro = 0
    output = ""
    for i in range(0,len(x)):
        chi_quadro += pow((y[i]-(b*x[i]+a)),2)/pow(errore_y[i],2)
    print("Il chi quadro vale:", chi_quadro)
    print("Il numero di DOF nel caso di 2 vincoli è: ", len(y)-2)
    output += f"Il chi quadro vale: {chi_quadro}"
    output += f"Il numero di DOF nel caso di 2 vincoli è: {len(y)-2}"
    if docs:
        return output
    return chi_quadro


def chi_squared_2(y, sigma_y ,x, a, b, c):
    chi_quadro = 0
    for i in range(len(y)):
        chi_quadro += ((y[i] - (a*x[i]**2 + b*x[i] +c))/sigma_y[i])**2
    print("Con D.O.F = ", len(y) - 3)
    return chi_quadro


# In[17]:


def errore_posteriori(b, a, y:list, x:list, docs:bool = False):
    somma = 0
    output = ""
    for i in range(0, len(y)):
        somma += pow(a+b*x[i]-y[i],2)
    errore_a_posteriori = somma/(len(y)-2)
    print("L'errore a posteriori dell'interpolazione vale: ", errore_a_posteriori)
    output += f"L'errore a posteriori dell'interpolazione vale: {errore_a_posteriori}"
    if docs:
        return output
    return errore_a_posteriori


# In[18]:


def t_Student(rho, N):
    errore_rho = m.sqrt((1-rho**2)/(N-2))
    t = rho/errore_rho
    print("La variabile di Student vale t =", t, "con ", N-2, "gradi di libertà")


# # Andamento temporale

# In[19]:


def andamento_temporale_relazione(b, a, y:list, errore_y, x:list, x_axis_name = "Occorrenze" ,colore = "#EF553B"):
    scarti = []
    #vincoli = N-2 == 9
    for i in range(0,len(y)):
        scarti.append((y[i]-(b*x[i]+a))/errore_y[i])
    layout = go.Layout(
        title= "",
        yaxis=dict(
                title = 'Scarti'
        ),
        xaxis=dict(
            title= str("$" + x_axis_name + "$")
        )
    )
    l = np.linspace(0,99,100)
    traccia = go.Scatter(
        x = l,
        y = scarti,
        mode = 'lines',
        name = 'Scarti',
        #y = 7.98x -1.1
        line = dict(
        #shape='spline',
#         color = str(input("Inserisci un colore (default #EF553B): "))
        color = str(colore)
                   ),
        )
#     print(traccia)   
#     print(type(traccia))
    l1 = np.linspace(0., len(y), 1000)
    retta = go.Scatter(
        x = l1,
        y = [0]*len(l1),
        mode = 'lines',
        name = "$y = 0$",
        line = dict(
            shape='spline',
            color = 'firebrick',
            width=5
            )
        )
   
    #fig = go.Figure(data=[traccia1, traccia2, traccia3, traccia4, traccia5, traccia6, traccia7, traccia8, traccia9, retta], layout = layout)
    fig = go.Figure(data=[traccia, retta], layout=layout)
    fig.show()


# In[20]:


def andamento_temporale(b, a, y:list, errore_y:list, x:list):
    scarti = []
    #vincoli = N-2 == 9
    for i in range(0,len(y)):
        scarti.append((y[i]-(b*x[i]+a))/errore_y[i])
    layout = go.Layout(
        title= "",
        yaxis=dict(
                title = 'Scarti'
        ),
        xaxis=dict(
            title= "x"
        )
    )
    l = np.linspace(0,99,100)
    traccia = go.Scatter(
        x = l,
        y = scarti,
        mode = 'lines',
        name = 'Scarti',
        #y = 7.98x -1.1
        line = dict(
        #shape='spline',
        color = '#EF553B'
                   ),
        )
    l1 = np.linspace(0., len(y), 1000)
    retta = go.Scatter(
        x = l1,
        y = [0]*len(l1),
        mode = 'lines',
        name = "$y = 0$",
        line = dict(
            shape='spline',
            color = 'firebrick',
            width=5
            )
        )
    fig = go.Figure(data=[traccia, retta], layout=layout)
    fig.show()


# In[21]:


def traccia(y, errore_y, x, errore_x = None, modo = 'markers'):
    #l = np.linspace(0,99,100)
    traccia1 = go.Scatter(
        x = x,
        y = y,
        mode = str(modo),
        name= "$x_i$",
        showlegend=True,
        error_y=dict(
                type='data',
                array = errore_y,
        ),
        error_x=dict(
                type='data',
                array = errore_x,
        ),
    )
    return traccia1

def retta(b, a, y, errore_y, x, errore_x = None):
    t_max, t_min = max(x), min(x)
    t = np.linspace(t_min, t_max, 10000) 
    lalla = a + b*t
    retta1 = go.Scatter(
        x = t,
        y = lalla,
        mode = 'lines',
        name = "Retta"
        #name = '$y = (25.594x - 0.02081) cm/s $',
#             line = dict(
#             shape='spline',
#             color = 'orange'
#        )  
    )

    return retta1

def retta_verticale(y, x, length = None):
    j = np.linspace(min(y), max(y), 10000)
    if x != list(x):
        t = [x]*length
    else: 
        t = x
    retta1 = go.Scatter(
        x = t,
        y = j,
        mode = 'lines',
        name = "Retta"
        #name = '$y = (25.594x - 0.02081) cm/s $',
#             line = dict(
#             shape='spline',
#             color = 'orange'
#        )  
    )
    return retta1


def parabola(a, b, c, y, x):
    t_max, t_min = max(x), min(x)
    t = np.linspace(t_min, t_max, 1000) 
    lalla = a*t**2 + b*t + c
    retta1 = go.Scatter(
        x = t,
        y = lalla,
        mode = 'lines',
        name = "Parabola"
        #name = '$y = (25.594x - 0.02081) cm/s $',
#             line = dict(
#             shape='spline',
#             color = 'orange'
#        )  
    )

    return retta1

def cubica(a, b, c, d, y, x):
    t_max, t_min = max(x), min(x)
    t = np.linspace(t_min, t_max, 1000) 
    lalla = a*t**3 + b*t**2 + c*t + d
    retta1 = go.Scatter(
        x = t,
        y = lalla,
        mode = 'lines',
        name = "Cubica"
        #name = '$y = (25.594x - 0.02081) cm/s $',
#             line = dict(
#             shape='spline',
#             color = 'orange'
#        )  
    )

    return retta1


def quadrata(a, b, c, d, e, y, x):
    t_max, t_min = max(x), min(x)
    t = np.linspace(t_min, t_max, 1000) 
    lalla = a*t**4 + b*t**3 + c*t**2 + d*t + e
    retta1 = go.Scatter(
        x = t,
        y = lalla,
        mode = 'lines',
        name = "Quadrata"
        #name = '$y = (25.594x - 0.02081) cm/s $',
#             line = dict(
#             shape='spline',
#             color = 'orange'
#        )  
    )

    return retta1


def Layout():
    layout = go.Layout(
        title= "",
        yaxis=dict(
                title = ""
        ),
        xaxis=dict(
            title= "",
        )
    )
    return layout
    


# In[ ]:





# In[22]:


def traccia_relazione(y, errore_y, x, errore_x = None, modo = 'markers'):
    #l = np.linspace(0,99,100)
    traccia1 = go.Scatter(
        x = x,
        y = y,
        mode = str(modo),
        name= "$" + str(input("Inserisci la legenda della traccia ") ) +"$",
#          name= str(input("Inserisci la legenda della traccia ") ) ,
        showlegend=True,
        error_y=dict(
                type='data',
                array = errore_y,
        ),
        error_x=dict(
                type='data',
                array = errore_x,
        ),
    )
    return traccia1

def retta_relazione(b:float, a:float, y, errore_y, x, errore_x = None):
#     i = int(input("Scegli di quanto vuoi arrotondare b: ") )
#     j = int(input("Scegli di quanto vuoi arrotondare a: ") )
    t_max, t_min = max(x), min(x)
    t = np.linspace(t_min, t_max, 1000) 
    lalla = a + b*t
    if a >= 0:
        retta1 = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = str("$y = " + "("+str(b)+ "x +" + str(a) +")" + str( input("Inserisci l'unità di misura della legenda: "))+"$")
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    else:
        retta1 = go.Scatter(
            x = t,
            y = lalla,
            mode = 'lines',
            name = str("$y = " + "("+str(b)+ "x" + str(a) +")" + str( input("Inserisci l'unità di misura della legenda: "))+"$")
            #name = '$y = (25.594x - 0.02081) cm/s $',
    #             line = dict(
    #             shape='spline',
    #             color = 'orange'
    #        )  
        )
    return retta1

def Layout_relazione():
    layout = go.Layout(
        title= str(input("Inserisci il nome del grafico ") ),
        yaxis=dict(
                title = str("$"+str(input("Inserisci il titolo dell'asse y ") ) + " [" +str(input("Inserisci l'unità di misura ")+ "]" + "$"))
        ),
        xaxis=dict(
            title= str("$" + str(input("Inserisci il titolo dell'asse x ") ) + " [" +str(input("Inserisci l'unità di misura "))+ "]" +"$"),
        )
    )
    return layout
    
def Layout_relazione_no_units():
    layout = go.Layout(
        title= str(input("Inserisci il nome del grafico ") ),
        yaxis=dict(
#                 title = str("$"+str(input("Inserisci il titolo dell'asse y ") ) + " [" +str(input("Inserisci l'unità di misura ")+ "]" + "$"))
                title = str("$"+str(input("Inserisci il titolo dell'asse y ") ) + "$")
        ),
        xaxis=dict(
            title= str("$" + str(input("Inserisci il titolo dell'asse x ") ) + "$"),
        )
    )
    return layout
    


# In[23]:


def Grafico(x:list, layout, log_x = False):
    fig = go.Figure(data=x, layout=layout)
#     fig.show()
    if log_x:
        fig.update_xaxes(type="log")
    fig.show()


# # Fit Lineare

# # $y = bx + a$

# In[24]:


def f(B, x):
    '''Linear function y = m*x + b'''
    # B is a vector of the parameters.
    # x is an array of the current x values.
    # x is in the same format as the x passed to Data or RealData.
    #
    # Return an array in the same format as y passed to Data or RealData.
    return B[0]*x + B[1]
def linear_fit(x, y, sigma_x, sigma_y, b, a):
    mydata = RealData(x, y, sx=sigma_x, sy=sigma_y)
    linear = Model(f)
    myodr = ODR(mydata, linear, beta0=[b, a])
    myoutput = myodr.run()
    myoutput.pprint()
    covariana_matrice = np.sqrt(np.diag(myoutput.cov_beta))
    return myoutput.beta[0], covariana_matrice[0], myoutput.beta[1], covariana_matrice[1]


# # $y = bx $ : $a \equiv 0$

# In[25]:


def f_1(B, x):
    '''Linear function y = m*x + b'''
    # B is a vector of the parameters.
    # x is an array of the current x values.
    # x is in the same format as the x passed to Data or RealData.
    #
    # Return an array in the same format as y passed to Data or RealData.
    return B[0]*x
def linear_fit(x, y, sigma_x, sigma_y, b):
    mydata = RealData(x, y, sx=sigma_x, sy=sigma_y)
    linear = Model(f_1)
    myodr = ODR(mydata, linear, beta0=[b])
    myoutput = myodr.run()
    myoutput.pprint()
    covariana_matrice = np.sqrt(np.diag(myoutput.cov_beta))
    return myoutput.beta, covariana_matrice


# # Fit Parabolico

# # $y = ax^2 + bx + c$

# In[26]:


def g(B, x):
    '''quadratic function y = a*x**2 + b*x + c'''
    # B is a vector of the parameters.
    # x is an array of the current x values.
    # x is in the same format as the x passed to Data or RealData.
    #
    # Return an array in the same format as y passed to Data or RealData.
    return B[0]*x**2 + B[1]*x + B[2]
def parabolic_fit(x, y, sigma_x, sigma_y, a, b, c):
    mydata = RealData(x, y, sx=sigma_x, sy=sigma_y)
    quadratic = Model(g)
    myodr = ODR(mydata, quadratic, beta0=[a, b, c])
    myoutput = myodr.run()
    myoutput.pprint()
    covariana_matrice = myoutput.cov_beta
    sigma_diagonale = np.sqrt(np.diag(myoutput.cov_beta))
#     return myoutput.beta[0], myoutput.sd_beta[0], myoutput.beta[1], myoutput.sd_beta[1], myoutput.beta[2], myoutput.sd_beta[2]
    return myoutput.beta[0], sigma_diagonale[0], myoutput.beta[1], sigma_diagonale[1], myoutput.beta[2], sigma_diagonale[2], covariana_matrice[0][1]
    


# In[27]:


def max_quadratic(a, b, c):
    x_max = -b/(2*a)
    y_max = a*x_max**2 + b*x_max +c
    return (y_max, x_max)


# In[28]:


def sigma_max_quadratic(a, sigma_a, b, sigma_b, c, sigma_c, covarianza_a_b):
    sigma_x = m.sqrt(((-b/(2*a))**2)*((sigma_a/a)**2 + (sigma_b/b)**2) - 2*b/(4*a**3)*covarianza_a_b)
    sigma_y = 0 # ToDo
#     var_x = (((-b/(2*a))**2)*((sigma_a/a)**2 + (sigma_b/b)**2) - 2*b/(4*a**3)*covarianza_a_b)
    return sigma_y, sigma_x


# In[ ]:





# # Fit Polinomiale

# In[29]:


def polinomial_fit(x, y, sigma_x, sigma_y, parametri, g):
    mydata = RealData(x, y, sx=sigma_x, sy=sigma_y)
    quadratic = Model(g)
    myodr = ODR(mydata, quadratic, beta0=parametri)
    myoutput = myodr.run()
    myoutput.pprint()
    return myoutput.beta, np.sqrt(np.diag(myoutput.cov_beta))


# In[30]:


# def max_polinomial():
#     pass


# In[31]:


from sympy import *
def sigma_max_polinomial(parametro1, sigma_parametro1, parametro2, sigma_parametro2, parametro3, sigma_parametro3, parametro4, sigma_parametro4, covarianza_matrice):
    x, a, b, c, d = sp.symbols('x, a, b, c, d', real=True) 
    q = d/a -(b*c)/(3*a**2) + (2*b**3)/(27*a**3)
    p = c/a - (b**2)/(3*a**2)
    x = -b/(3*a) + (-(d/a -(b*c)/(3*a**2) + (2*b**3)/(27*a**3))/2 + ((((d/a -(b*c)/(3*a**2) + (2*b**3)/(27*a**3))**2)/4 + (c/a - (b**2)/(3*a**2))**3/27))**(1/2))**(1/3) + (-(d/a -(b*c)/(3*a**2) + (2*b**3)/(27*a**3))/2 - ((((d/a -(b*c)/(3*a**2) + (2*b**3)/(27*a**3))**2)/4 + (c/a - (b**2)/(3*a**2))**3/27))**(0.5))**(1/3)
    x_prime_a = x.diff(a)
    x_prime_b = x.diff(b)
    x_prime_c = x.diff(c)
    x_prime_d = x.diff(d)
    x_prime_a_value = x_prime_a.subs([(a, parametro1), (b, parametro2), (c, parametro3), (d, parametro4)])
    x_prime_b_value = x_prime_b.subs([(a, parametro1), (b, parametro2), (c, parametro3), (d, parametro4)])
    x_prime_c_value = x_prime_c.subs([(a, parametro1), (b, parametro2), (c, parametro3), (d, parametro4)])
    x_prime_d_value = x_prime_d.subs([(a, parametro1), (b, parametro2), (c, parametro3), (d, parametro4)])
    
    var_x = 0
    
    var_x += ((x_prime_a_value)*sigma_parametro1)**2
    var_x += ((x_prime_b_value)*sigma_parametro2)**2
    var_x += ((x_prime_c_value)*sigma_parametro3)**2
    var_x += ((x_prime_d_value)*sigma_parametro4)**2
    var_x += 2*x_prime_a_value*x_prime_b_value*covarianza_matrice[0][1]
    var_x += 2*x_prime_a_value*x_prime_c_value*covarianza_matrice[0][2]
    var_x += 2*x_prime_a_value*x_prime_d_value*covarianza_matrice[0][3]
    var_x += 2*x_prime_b_value*x_prime_c_value*covarianza_matrice[1][2]
    var_x += 2*x_prime_b_value*x_prime_d_value*covarianza_matrice[1][3]
    var_x += 2*x_prime_c_value*x_prime_d_value*covarianza_matrice[2][3]
#     print("La varianza complessa vale:", N(var_x))
    sigma_x = var_x**(0.5)
#     print("La deviazione standard complessa vale:", N(sigma_x))
    return N((re(sigma_x)**2 + im(sigma_x)**2)**(0.5))


# In[ ]:





# # Ricerca del minimo/massimo di un fit parabolico

# In[32]:


def ricerca_minimo(y, sigma_y, x): 
    #Nota Bene: le derivate vengono calcolate per n-2 punti, quindi gliene servono almeno 5 in entrata (ma gli estremi non sono considerati). 
    #Funziona per interpolazione della derivata, per cui se la relazione non è parabolica dà risultati approssimati
    d=[]
    d_x=[]
    h=(x[1]-x[0]) 
    sigma_d=2**0.5*sigma_y/h

    for i in range(1,len(y)-1):
        
        this_d=(y[i+1]-y[i-1])/h  #Calcolo delle derivate con il metodo delle differenze finite
        d.append(this_d)
        d_x.append(x[i])
        
    N = len(d)
    x_quadri, x_singoli, d_singoli, x_per_d = 0, 0, 0, 0
    for i in range (0,len(d)):
        x_quadri += pow(d_x[i],2)
        x_singoli += d_x[i]
        d_singoli += d[i]
        x_per_d += d[i]*d_x[i]

    delta = N*x_quadri -pow(x_singoli,2)
    a = (x_quadri*d_singoli-x_singoli*x_per_d)/delta
    b = (N*x_per_d-x_singoli*d_singoli)/delta
    errore_a = sigma_d*(x_quadri/delta)**0.5
    errore_b = sigma_d*(N/delta)**0.5
    x_min=(-a/b)
    err_x=((errore_a/a)**2+(errore_b/b)**2)**0.5/x_min
    output = 'Il minimo ha coordinata x ' + str(x_min) + '\ned errore ' + str(err_x)
    return (output)
    
    
    


# # Scrittura su file

# In[33]:


def scrittura(nome_file, contenuto):
    ciao = nome_file
    f = open(ciao, "w")
    f.write(contenuto)
    #f.read()
    #f.readlines()
    f.close()
    f = open(nome_file, "a")
    a_capo = "\n"
    f.write(a_capo)
    f.close()
def scrittura_aggiunta(nome_file, contenuto):
    f = open(nome_file, "a")
    f.write(contenuto)
    a_capo = "\n"
    f.write(a_capo)
    f.close()


# # Importare da Excel

# In[34]:


import openpyxl
def excel_import(nome_file:str, start, stop, col):
    #excel_document = openpyxl.load_workbook('Guidovia.xlsx')
    #sheet = excel_document.get_sheet_by_name('Foglio1')
    nome_file += '.xlsx'
    wb = openpyxl.load_workbook(nome_file)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    vettore = [] #array contenete le medie delle misure effettuate della variabile diretta.
    for i in range(start, stop + 1):
        casella = ws.cell(row=i, column=col)
        vettore.append(casella.value) #aggiunge elementi al vettore (array)
    return vettore


# # Leggi file

# In[35]:


def text(nome_file:str):
    h = open(nome_file, 'r')
    # Reading from the file
    content = h.readlines()
    # print(content)
    h.close()
    n = 0
    x = []
    y = []
    for line in content:
        if n==0:
            N = line[0]
        if n>0:
            words = line.split()
            if len(words):
                x.append(words[0])
                y.append(words[1])
        n+=1
    return(x,y)

